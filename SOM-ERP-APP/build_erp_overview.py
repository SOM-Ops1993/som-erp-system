"""
SOM ERP System — MD Overview Workbook
8 sheets covering every layer of the system
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ──────────────────────────────────────────────────────────────────
DARK_GREEN  = "1A4731"
MID_GREEN   = "2D6A4F"
LIGHT_GREEN = "D8F3DC"
ACCENT_GREEN= "52B788"
HEADER_GRAY = "F8F9FA"
BORDER_GRAY = "DEE2E6"
AMBER       = "FFF3CD"
AMBER_DARK  = "856404"
BLUE_LIGHT  = "E8F4FD"
BLUE_DARK   = "0D6EFD"
PURPLE_LIGHT= "F3E8FF"
PURPLE_DARK = "6F42C1"
RED_LIGHT   = "FFF0F0"
RED_DARK    = "DC3545"
ORANGE_LIGHT= "FFF4E6"
ORANGE_DARK = "E55A00"
WHITE       = "FFFFFF"
BLACK       = "000000"
GRAY_TEXT   = "6C757D"

def side(color=BORDER_GRAY, style="thin"):
    return Side(border_style=style, color=color)

def border(color=BORDER_GRAY, style="thin"):
    s = side(color, style)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def font(bold=False, size=11, color=BLACK, italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def hdr(ws, row, col, text, bg=DARK_GREEN, fg=WHITE, sz=11, bold=True,
        h="center", span=1, italic=False):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(bold=bold, size=sz, color=fg, italic=italic)
    c.alignment = align(h=h)
    c.border = border(bg)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return c

def cell(ws, row, col, value="", bg=WHITE, fg=BLACK, bold=False,
         h="left", wrap=True, italic=False, sz=10):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=bold, size=sz, color=fg, italic=italic)
    c.alignment = align(h=h, wrap=wrap)
    c.border = border()
    return c

def section_title(ws, row, col, text, span, bg=MID_GREEN, fg=WHITE, sz=11):
    c = hdr(ws, row, col, text, bg=bg, fg=fg, sz=sz, span=span, h="left")
    return c

def banner(ws, row, text, ncols, bg=DARK_GREEN, fg=WHITE, sz=14):
    ws.row_dimensions[row].height = 30
    c = hdr(ws, row, 1, text, bg=bg, fg=fg, sz=sz, span=ncols, h="left", bold=True)
    return c

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MASTERS
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("1. Masters")
set_col_widths(ws, [22, 28, 32, 28, 22, 22, 22, 22])
ws.freeze_panes = "A3"

banner(ws, 1, "  SOM ERP — Master Data Overview", 8)
ws.row_dimensions[2].height = 6

MASTERS = [
    {
        "name": "Product Master",
        "desc": "Central catalog of all finished goods / SFG the company manufactures.",
        "key": "productCode (PROD-001…)",
        "cols": ["productCode", "productName", "sectionName", "uom", "shelfLife (days)", "isActive"],
        "notes": "Every BOM, production plan, and batch sheet links back to this master via productCode.",
        "bg": LIGHT_GREEN, "hbg": MID_GREEN,
    },
    {
        "name": "Raw Material (RM) Master",
        "desc": "All raw materials, packaging materials, and consumables used in production.",
        "key": "itemCode (RM-001…)",
        "cols": ["itemCode", "itemName", "category", "uom", "reorderLevel", "leadTimeDays"],
        "notes": "RM Master drives BOM line items. Stock levels compared here for RM Availability check in planning.",
        "bg": BLUE_LIGHT, "hbg": BLUE_DARK,
    },
    {
        "name": "Customer Master",
        "desc": "All customers: name, company, order type defaults.",
        "key": "customerName (unique)",
        "cols": ["customerName", "company", "orderType", "city", "state", "gstNo"],
        "notes": "Linked to every Sales Order. customerName drives Customer Product Profile lookup.",
        "bg": AMBER, "hbg": AMBER_DARK,
    },
    {
        "name": "Customer Product Profile",
        "desc": "Memory of every product ordered by each customer — packing, specs, labels, MRP, batch pattern. 2100+ profiles seeded.",
        "key": "(customerName, productName)",
        "cols": ["customerName", "productName", "productCode", "inhouseName", "unitQty+UOM",
                 "primaryPack", "secondaryPack", "labelType", "mrp", "activeSpecs", "carrier", "sectionName"],
        "notes": "Auto-fills Sales Order line when customer product is selected. Gets richer with every order.",
        "bg": PURPLE_LIGHT, "hbg": PURPLE_DARK,
    },
    {
        "name": "Equipment Master",
        "desc": "All production equipment across plants (NANO, BOTANICAL, LIQUID, POWDER, GRANULES).",
        "key": "equipmentId",
        "cols": ["equipmentId", "name", "plant/section", "capacity (KG)", "status", "lastService"],
        "notes": "Planning engine checks equipment availability before scheduling. Busy/Available/Unknown status shown on plan card.",
        "bg": ORANGE_LIGHT, "hbg": ORANGE_DARK,
    },
    {
        "name": "Recipe / BOM Master",
        "desc": "Bill of Materials — lists all RMs and their quantities per KG of finished product.",
        "key": "(productCode, rmCode)",
        "cols": ["productCode", "productName", "rmCode", "rmName", "qtyPerUnit", "uom", "roleType"],
        "notes": "roleType: ACTIVE / CARRIER / EXCIPIENT / PACKING. Formulation BOM = ACTIVE+CARRIER+EXCIPIENT; Packing BOM = PACKING.",
        "bg": LIGHT_GREEN, "hbg": ACCENT_GREEN,
    },
    {
        "name": "Employee Master",
        "desc": "All employees — used for batch incharge assignment and shift planning.",
        "key": "employeeId",
        "cols": ["employeeId", "name", "department", "designation", "shift", "section"],
        "notes": "Batch incharge on a production plan is selected from this master.",
        "bg": RED_LIGHT, "hbg": RED_DARK,
    },
    {
        "name": "Company Master",
        "desc": "All company codes under the group — DVS, AML, etc.",
        "key": "companyCode",
        "cols": ["companyCode", "companyName", "address", "gstNo"],
        "notes": "Sales Orders are tagged to a company. Used for dispatch documentation.",
        "bg": BLUE_LIGHT, "hbg": BLUE_DARK,
    },
]

r = 3
for m in MASTERS:
    ws.row_dimensions[r].height = 20
    section_title(ws, r, 1, f"  {m['name']}", span=8, bg=m["hbg"], fg=WHITE, sz=11)
    r += 1

    ws.row_dimensions[r].height = 44
    cell(ws, r, 1, "Description", bg=HEADER_GRAY, bold=True, sz=10)
    cell(ws, r, 2, m["desc"], bg=HEADER_GRAY, sz=10, h="left", wrap=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    cell(ws, r, 5, "Primary Key", bg=HEADER_GRAY, bold=True, sz=10)
    cell(ws, r, 6, m["key"], bg=HEADER_GRAY, sz=10, italic=True)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    r += 1

    cols = m["cols"]
    for ci, col in enumerate(cols[:8], 1):
        hdr(ws, r, ci, col, bg=m["hbg"], fg=WHITE, sz=9)
    r += 1

    cell(ws, r, 1, "(populated from actual database — see system for live data)", bg=m["bg"],
         italic=True, fg=GRAY_TEXT, sz=9)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

    ws.row_dimensions[r].height = 30
    cell(ws, r, 1, "Notes:", bold=True, sz=9, bg=HEADER_GRAY)
    cell(ws, r, 2, m["notes"], sz=9, bg=HEADER_GRAY, wrap=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1

    ws.row_dimensions[r].height = 8
    r += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — PROCESS FLOW
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Process Flow")
set_col_widths(ws2, [4, 26, 34, 34, 22, 20, 24])
ws2.freeze_panes = "A4"

banner(ws2, 1, "  SOM ERP — End-to-End Production Process (Step by Step)", 7)
ws2.row_dimensions[2].height = 6

for ci, h_text in enumerate(["#", "Stage", "What Happens", "System Actions",
                              "Who Acts", "Status After", "Module"], 1):
    hdr(ws2, 3, ci, h_text, bg=DARK_GREEN, sz=10)

PROCESS = [
    ("ORDER CREATION", "", "", "", "", ""),
    ("1", "Sales Order Entry",
     "Sales team enters DI No., Customer, Products, Qty, Batch details, Packing specs, Label type, MRP, Mfg/Exp dates.",
     "Auto-fills packing/specs from Customer Product Profile (2100+ known products). Validates all fields. Saves to DB.",
     "Sales Team", "PENDING", "Sales Orders"),
    ("2", "SFG Check on Save",
     "When order is saved, system checks if any Semi-Finished Goods (SFG) stock already exists for the ordered product.",
     "Backend queries SFG table by productCode. If SFG found, planner gets alert with SFG qty and batch details.",
     "System (auto)", "PENDING", "Sales Orders + SFG"),
    ("PLANNING PHASE", "", "", "", "", ""),
    ("3", "Plan Engine Run",
     "Planner clicks Run Planning Engine, or it auto-runs at 08:30 daily. Picks all PENDING order items.",
     "Resolves productCode from Product Master. Checks RM availability vs BOM. Sets planned date (ETD minus 3 days). Creates Production Plan cards.",
     "Planner / Auto", "DRAFT Plan", "Planning Engine"),
    ("4", "Plan Review",
     "Planner reviews each plan card. Sees RM status (Green/Amber/Red), equipment availability, SFG alert if applicable.",
     "Plan card shows RM shortfall details, equipment status, SFG qty. Planner edits equipment, shift, batch size, cycles.",
     "Planner", "REVIEWED", "Planning"),
    ("5", "Send to Schedule",
     "Planner confirms plan and sends to schedule. System raises Production Indent and creates batch entry in Production Master.",
     "1) Indent created — RM qty x cycles from BOM. 2) Batch entry in Production Master. 3) Plan status = IN_PROGRESS.",
     "Planner", "IN_PROGRESS", "Planning + Indent + Prod Master"),
    ("FORMULATION PHASE", "", "", "", "", ""),
    ("6", "BOM Issuance (Formulation)",
     "Store issues raw materials against the formulation BOM. QR scan confirms each RM batch issued.",
     "BOM Send (type=FORMULATION). Store scans RM QR codes. Stock ledger debited. Issuance record created.",
     "Store + Production", "RM Issued", "BOM Issuance + Store"),
    ("7", "Formulation Batch Sheet (BMR)",
     "Production fills digital BMR — temperatures, observations, yield, CFU count, operator sign-off.",
     "BMR pre-filled with product specs and BOM quantities. Operator enters actuals. System flags deviations.",
     "Production Operator", "Formulated", "Batch Sheet"),
    ("8", "SFG Storage (Optional)",
     "If formulated material is stored rather than packed immediately, it is entered as SFG stock.",
     "SFG entry created: product, batch no, qty, formulation date, location. Future orders see this SFG.",
     "Production + Store", "SFG Stored", "SFG Management"),
    ("PACKING PHASE", "", "", "", "", ""),
    ("9", "BOM Issuance (Packing)",
     "Store issues packing materials against the packing BOM — pouches, boxes, labels, seals.",
     "BOM Send (type=PACKING). PM debited from stock. Issuance record linked to batch.",
     "Store + Packing", "PM Issued", "BOM Issuance + Store"),
    ("10", "Packing Batch Sheet (BMR)",
     "Packing team fills digital packing BMR — pack count, weight checks, seal integrity, label verification.",
     "Pre-filled with packing specs from plan. Operator enters actuals. Linked to formulation batch via batchNo.",
     "Packing Operator", "Packed", "Batch Sheet"),
    ("LABELING & FINISHING", "", "", "", "", ""),
    ("11", "Labeling",
     "Labels applied with MRP, Mfg Date, Exp Date, Batch No. Label type pre-specified in order.",
     "Label specs fetched from plan. QR label may be generated for each finished pack.",
     "Packing + Finishing", "Labeled", "Batch Sheet + Labels"),
    ("12", "FG Inward",
     "Packed and labeled goods entered into FG store with QR location tagging.",
     "FG stock created. Each pack assigned a QR. Location tagged. Stock ledger credited.",
     "Store", "FG In Stock", "Store + QR System"),
    ("DISPATCH PHASE", "", "", "", "", ""),
    ("13", "Dispatch Planning",
     "Dispatch team matches FG stock to sales order. Raises dispatch instruction.",
     "System verifies FG qty vs order qty. Generates dispatch note with DI No., Customer, Product, Batch, Qty.",
     "Dispatch + Sales", "Ready to Ship", "Dispatch"),
    ("14", "Gate Pass and Outward",
     "Security gate records vehicle, driver, seal no. Outward entry created.",
     "Gate pass generated. Stock ledger debited (outward). Delivery challan / invoice linked.",
     "Gate + Store", "DISPATCHED", "Gate + Outward"),
    ("15", "Order Closure",
     "Sales order marked DISPATCHED. Customer informed. Records archived.",
     "SalesOrder status = DISPATCHED. All linked records closed.",
     "Sales + System", "CLOSED", "Sales Orders"),
]

r = 4
for row in PROCESS:
    if row[1] == "":
        ws2.row_dimensions[r].height = 22
        section_title(ws2, r, 1, f"  {row[0]}", span=7, bg=ACCENT_GREEN, fg=WHITE, sz=10)
        r += 1
        continue
    ws2.row_dimensions[r].height = 55
    step_no, stage, what, sys_action, who, status, module = row
    cell(ws2, r, 1, step_no, bg=LIGHT_GREEN, bold=True, h="center", sz=10)
    cell(ws2, r, 2, stage, bold=True, sz=10, bg=HEADER_GRAY)
    cell(ws2, r, 3, what, sz=9, wrap=True)
    cell(ws2, r, 4, sys_action, sz=9, wrap=True, bg=BLUE_LIGHT)
    cell(ws2, r, 5, who, sz=9, h="center")
    s_bg = LIGHT_GREEN if status in ("CLOSED", "DISPATCHED") else \
           AMBER if status in ("DRAFT Plan", "PENDING") else BLUE_LIGHT
    cell(ws2, r, 6, status, sz=9, h="center", bg=s_bg, bold=True)
    cell(ws2, r, 7, module, sz=9, italic=True, fg=MID_GREEN)
    r += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — INTERCONNECTIONS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Interconnections")
set_col_widths(ws3, [26, 30, 30, 26, 20, 22])
ws3.freeze_panes = "A4"

banner(ws3, 1, "  SOM ERP — Event Trigger Map (What happens when something happens)", 6)
ws3.row_dimensions[2].height = 6

for ci, h_text in enumerate(["Event / Trigger", "Immediate System Actions",
                              "Downstream Modules Affected",
                              "Data Written / Updated", "Alert Raised?", "Who Sees It"], 1):
    hdr(ws3, 3, ci, h_text, bg=DARK_GREEN, sz=10)

INTERCONNECTIONS = [
    ("Sales Order Saved",
     "1. Customer Profile upserted\n2. SFG stock checked per line item\n3. Order items set PENDING",
     "Customer Product Profiles, SFG, Notifications",
     "SalesOrder, SalesOrderItem, CustomerProductProfile",
     "Yes — if SFG found", "Sales Team, Planner"),
    ("Plan Engine Runs",
     "1. PENDING items fetched\n2. productCode resolved from Product Master\n3. RM availability checked vs BOM\n4. Production Plan created per item",
     "ProductionPlan, Equipment, RM Stock",
     "ProductionPlan rows, PlanSequence counter",
     "Yes — RM shortfall alerts", "Planner"),
    ("Planner Sends to Schedule",
     "1. Production Indent created (RM qty x cycles)\n2. Production Master batch entry created\n3. Plan = IN_PROGRESS\n4. SO Item = UNDER_PRODUCTION",
     "Indent Management, Production Master, SalesOrderItem",
     "ProductionIndent, ProductionBatch, plan.status, soItem.status",
     "Yes — batch created notification", "Store, Production"),
    ("BOM Send (Formulation)",
     "1. Formulation BOM lines fetched from RecipeDb\n2. BOM Send record created\n3. Store receives issuance task",
     "BOM Issuance, Store, Stock Ledger",
     "BomSend, BomIssuanceLine rows",
     "Yes — store alert to issue RMs", "Store Manager"),
    ("BOM Send (Packing)",
     "1. Packing material lines fetched (roleType=PACKING)\n2. BOM Send created\n3. Store issuance task raised",
     "BOM Issuance, Store, Stock Ledger",
     "BomSend (type=PACKING), BomIssuanceLine",
     "Yes — store alert", "Store Manager"),
    ("Store Issues RM (QR Scan)",
     "1. RM stock debited in Stock Ledger\n2. BOM Issuance line = ISSUED\n3. Running balance updated",
     "Stock Ledger, BOM Issuance",
     "StockLedger (debit), BomIssuanceLine.status=ISSUED",
     "No (unless short-issue)", "Production, Store"),
    ("SFG Stock Entry Created",
     "1. SFG record created with batch, qty, formulation date\n2. Future planning engine sees this SFG\n3. Planner alerted on next order for same product",
     "SFG Management, Planning Engine, Notifications",
     "MicrobialSfg / SfgStock record",
     "Yes — planner notified on next order", "Planner"),
    ("FG Inward (QR Scan)",
     "1. FG pack QR generated\n2. Location tagged\n3. Stock Ledger credited",
     "Store / QR System, Stock Ledger, FG Inventory",
     "FgPack, QrLocation, StockLedger (credit)",
     "No", "Store"),
    ("Dispatch / Outward",
     "1. FG stock debited\n2. Gate pass generated\n3. SalesOrder = DISPATCHED\n4. Outward record created",
     "Stock Ledger, Gate, SalesOrder, Outward",
     "StockLedger (debit), OutwardEntry, SalesOrder.status",
     "Yes — dispatch confirmation", "Sales, Finance"),
    ("New RM GRN",
     "1. RM stock credited\n2. Stock Ledger updated\n3. QR label printed for received lot",
     "Stock Ledger, RM Master, QR Labels",
     "StockLedger (credit), GrnEntry",
     "Yes — if pending indent now satisfied", "Store, Planner"),
    ("Direct BOM Issuance (No Recipe)",
     "Planner manually opens Indent Management. Adds RM/PM lines manually. Issues without a BOM recipe.",
     "Indent Management, Stock Ledger",
     "ProductionIndent (manual), BomIssuance",
     "No", "Planner, Store"),
]

r = 4
for row in INTERCONNECTIONS:
    ws3.row_dimensions[r].height = 60
    event, actions, downstream, data, alert, who = row
    cell(ws3, r, 1, event, bold=True, bg=LIGHT_GREEN, sz=10)
    cell(ws3, r, 2, actions, sz=9, wrap=True)
    cell(ws3, r, 3, downstream, sz=9, wrap=True, bg=BLUE_LIGHT)
    cell(ws3, r, 4, data, sz=9, wrap=True, italic=True, fg=GRAY_TEXT)
    a_bg = AMBER if "Yes" in alert else HEADER_GRAY
    cell(ws3, r, 5, alert, sz=9, h="center", bg=a_bg, bold=("Yes" in alert))
    cell(ws3, r, 6, who, sz=9, fg=MID_GREEN, italic=True)
    r += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — NOTIFICATIONS
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. Notifications")
set_col_widths(ws4, [28, 36, 22, 18, 20, 18])
ws4.freeze_panes = "A4"

banner(ws4, 1, "  SOM ERP — Notification Plan (All System Alerts & Digests)", 6)
ws4.row_dimensions[2].height = 6

for ci, h_text in enumerate(["Notification Event", "Message Content",
                              "Recipient(s)", "Channel", "Timing", "Status"], 1):
    hdr(ws4, 3, ci, h_text, bg=DARK_GREEN, sz=10)

NOTIFS = [
    ("SFG Available on Order Save",
     "SFG Alert: [Product] has [X KG] semi-finished stock in Batch [B-001]. Consider skipping formulation and planning packing directly.",
     "Planner", "In-App Banner", "On order save", "LIVE"),
    ("RM Shortfall in Plan",
     "RM Shortfall: [RM Name] requires [X KG] but only [Y KG] available for plan [PP-2025-0001].",
     "Planner, Store Manager", "In-App", "On plan creation", "LIVE"),
    ("Plan Ready for Review",
     "New production plan PP-XXXX created for [Product] — [Customer]. Please review and schedule.",
     "Planner", "In-App", "On engine run", "LIVE"),
    ("Production Indent Created",
     "Production Indent raised for [Product], Batch [B-XXX], [X KG] x [N] cycles. RM issuance required.",
     "Store Manager", "In-App", "On Send to Schedule", "LIVE"),
    ("BOM Issuance Task",
     "BOM issuance required: [Formulation/Packing] BOM for [Product] — [X] line items to issue.",
     "Store Team", "In-App", "On BOM Send", "LIVE"),
    ("Low Stock Warning",
     "RM [Name] stock below reorder level. [X KG] remaining, reorder at [Y KG]. Raise purchase order.",
     "Store Manager, Purchase", "In-App + Email", "Daily check", "PLANNED"),
    ("Order Overdue (ETD Passed)",
     "OVERDUE: Sales Order [DI-XXX] for [Customer] was due [Date]. Current status: [status]. Action required.",
     "Sales, MD", "In-App + WhatsApp", "Daily at 09:00", "PLANNED"),
    ("Batch Completed",
     "Batch [B-XXX] of [Product] completed. [X KG] ready for packing / dispatch.",
     "Production Manager, Sales", "In-App", "On batch close", "PLANNED"),
    ("Dispatch Confirmed",
     "Order [DI-XXX] dispatched to [Customer] — [X KG] [Product], Batch [B-XXX], Vehicle [XX-00-XX-0000].",
     "Sales, Customer (optional)", "In-App + WhatsApp", "On dispatch", "PLANNED"),
    ("GRN Received — Indent Match",
     "GRN: [X KG] of [RM Name] received. Pending production indent [Indent-XXX] can now be fulfilled.",
     "Planner, Store", "In-App", "On GRN save", "PLANNED"),
    ("Equipment Maintenance Due",
     "Equipment [Name] is due for service on [Date]. Schedule maintenance to avoid planning conflicts.",
     "Maintenance, Planner", "In-App + Email", "7 days before due date", "PLANNED"),
    ("Batch Expiry Warning",
     "SFG/RM Batch [B-XXX] expires on [Date] — [X days] remaining. Use immediately or arrange disposal.",
     "Store, Production", "In-App", "14 days before expiry", "PLANNED"),
    ("Daily MD Summary",
     "Daily Digest: [N] orders pending, [N] plans active, [N] batches in production, [N] dispatched today, [N] RM shortfalls.",
     "MD, Production Manager", "Email + WhatsApp", "Daily at 08:00", "PLANNED"),
]

r = 4
for row in NOTIFS:
    ws4.row_dimensions[r].height = 50
    event, msg, recip, channel, timing, status = row
    cell(ws4, r, 1, event, bold=True, bg=LIGHT_GREEN, sz=10)
    cell(ws4, r, 2, msg, sz=9, wrap=True)
    cell(ws4, r, 3, recip, sz=9, italic=True, fg=MID_GREEN)
    cell(ws4, r, 4, channel, sz=9, h="center")
    cell(ws4, r, 5, timing, sz=9, h="center")
    s_bg = LIGHT_GREEN if status == "LIVE" else AMBER
    cell(ws4, r, 6, status, sz=9, h="center", bg=s_bg, bold=True)
    r += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — FUTURE AUTOMATION
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5. Automation Roadmap")
set_col_widths(ws5, [26, 36, 30, 18, 16, 16])
ws5.freeze_panes = "A4"

banner(ws5, 1, "  SOM ERP — Automation Roadmap (Every Planned Automation, Phase by Phase)", 6)
ws5.row_dimensions[2].height = 6

for ci, h_text in enumerate(["Area", "What Gets Automated",
                              "How It Works (Technical)", "Priority", "Phase", "Status"], 1):
    hdr(ws5, 3, ci, h_text, bg=DARK_GREEN, sz=10)

AUTOMATIONS = [
    ("Planning Engine", "Auto-create production plans from confirmed sales orders at 08:30 daily.",
     "Scheduled cron job. Reads all PENDING order items. Resolves product codes. Checks BOM+RM. Creates plan cards.",
     "P0 Critical", "Phase 1", "LIVE"),
    ("Customer Product Memory", "Auto-fill all packing, spec, label, MRP fields on sales order entry.",
     "CustomerProductProfile table (2100+ records). On product selection, all fields pre-populated instantly.",
     "P0 Critical", "Phase 1", "LIVE"),
    ("SFG Alert on Order Save", "Alert planner instantly if semi-finished goods exist for a new order.",
     "POST /sales-orders backend checks SFG table for productCode match. Returns alert with qty and batch.",
     "P0 Critical", "Phase 1", "LIVE"),
    ("Product Code Auto-Resolve", "Plans with missing product code auto-resolve by name from Product Master.",
     "Plan engine + GET /plans both do case-insensitive name lookup and persist resolved code.",
     "P0 Critical", "Phase 1", "LIVE"),
    ("Batch No Suggestion", "Next batch number auto-suggested based on last batch pattern per product.",
     "System reads lastBatchNo from CustomerProductProfile. Increments numeric suffix. Planner can override.",
     "P1 High", "Phase 1", "LIVE"),
    ("RM Shortage Auto-Alert", "Auto-notify store + purchase when any RM falls below reorder level.",
     "Daily batch job: check stock ledger balance vs RM Master reorder level. Create in-app notification.",
     "P1 High", "Phase 2", "PLANNED"),
    ("Expiry Auto-Watch", "Flag RMs and SFG batches approaching expiry (14 days out, escalate at 7).",
     "Daily job scans all lots with expDate. Raises alert at 14 days. Escalation at 7 days.",
     "P1 High", "Phase 2", "PLANNED"),
    ("Order Overdue Escalation", "Auto-escalate orders crossing ETD without dispatch.",
     "Nightly: if ETD < today and status != DISPATCHED, create escalation notification for MD.",
     "P1 High", "Phase 2", "PLANNED"),
    ("Daily MD Summary Email/WhatsApp", "Automated 08:00 digest to MD with full plant status.",
     "Scheduled job aggregates dashboard metrics. Sends formatted email + WhatsApp via API.",
     "P1 High", "Phase 2", "PLANNED"),
    ("Auto Purchase Indent", "When RM falls below reorder level, auto-raise purchase indent.",
     "RM stock watch triggers auto-indent with standard vendor, qty = reorder qty.",
     "P2 Medium", "Phase 3", "FUTURE"),
    ("WhatsApp Order Confirmation", "Customer receives WhatsApp confirmation when order is created.",
     "Webhook to WhatsApp Business API sends order summary to customer contact on save.",
     "P2 Medium", "Phase 3", "FUTURE"),
    ("Digital BMR e-Sign", "Batch incharge signs digital batch sheet from mobile — no paper needed.",
     "BMR UI has OTP-based e-sign. Locked post-sign. PDF auto-generated for records.",
     "P2 Medium", "Phase 3", "FUTURE"),
    ("Equipment OEE Tracking", "Track Overall Equipment Effectiveness per equipment per month.",
     "Batch sheet captures start/stop times. OEE = Availability x Performance x Quality.",
     "P2 Medium", "Phase 3", "FUTURE"),
    ("Lab Results Integration", "CFU count, assay results entered in batch sheet. Auto-compared to spec.",
     "Batch sheet lab section. Pass/Fail auto-determined. OOS triggers QA hold alert.",
     "P2 Medium", "Phase 3", "FUTURE"),
    ("Customer Order Tracking Portal", "Customer logs in to see their order status in real time.",
     "Read-only portal: DI No input. Shows plan status, batch status, expected dispatch.",
     "P3 Low", "Phase 4", "FUTURE"),
    ("AI Demand Forecasting", "System suggests next production run based on order history patterns.",
     "ML model on historical sales. Recommends pre-production qty by product by month.",
     "P3 Low", "Phase 5", "FUTURE"),
]

r = 4
for row in AUTOMATIONS:
    ws5.row_dimensions[r].height = 55
    area, what, how, priority, phase, status = row
    cell(ws5, r, 1, area, bold=True, bg=LIGHT_GREEN, sz=10)
    cell(ws5, r, 2, what, sz=9, wrap=True)
    cell(ws5, r, 3, how, sz=9, wrap=True, bg=BLUE_LIGHT)
    p_bg = RED_LIGHT if "P0" in priority else AMBER if "P1" in priority else HEADER_GRAY
    cell(ws5, r, 4, priority, sz=9, h="center", bg=p_bg, bold=True)
    cell(ws5, r, 5, phase, sz=9, h="center")
    s_bg = LIGHT_GREEN if status == "LIVE" else AMBER if status == "PLANNED" else PURPLE_LIGHT
    cell(ws5, r, 6, status, sz=9, h="center", bg=s_bg, bold=True)
    r += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — BATCH SHEET DIGITALIZATION
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6. Batch Sheet Digital")
set_col_widths(ws6, [24, 26, 34, 28, 22, 16])
ws6.freeze_panes = "A5"

banner(ws6, 1, "  SOM ERP — Batch Sheet Digitalization Plan (BMR for Every Plant & Product)", 6)
ws6.row_dimensions[2].height = 6

c = ws6.cell(row=3, column=1,
             value="OBJECTIVE: Every batch of every product in every plant has a digital Batch Manufacturing Record (BMR). "
                   "System pre-fills what it knows (product, batch, BOM, specs). Operator fills what happened. "
                   "Record locked after sign-off. Printable PDF template for physical filing and regulatory purposes.")
c.fill = fill(LIGHT_GREEN)
c.font = font(size=10, italic=True)
c.alignment = align(wrap=True)
c.border = border()
ws6.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
ws6.row_dimensions[3].height = 44

for ci, h_text in enumerate(["BMR Type", "Plant / Section",
                              "System Pre-Fills (Auto)",
                              "Operator Enters (Manual)", "Linked To", "Printable?"], 1):
    hdr(ws6, 4, ci, h_text, bg=DARK_GREEN, sz=10)

BMR_TYPES = [
    ("Formulation BMR",
     "NANO / BOTANICAL / LIQUID / POWDER / GRANULES",
     "Product name, Product code, Batch No, Batch Size (KG), Planned date, BOM qty for each RM, Carrier, Active specs, Operator name, Equipment",
     "Actual RM lot nos. used, Actual weights, Water / Solvent qty, Process observations, Temperature log, Yield obtained (KG), CFU count / Assay result, Remarks, Start and end time",
     "ProductionPlan, BOM Issuance, SFG (if stored)",
     "YES — PDF for BMR binder"),
    ("Packing BMR",
     "All sections (Packing area)",
     "Product name, Batch No, Pack size (unit qty + UOM), Primary + secondary packing type, Units per case, Total cases, Label type, MRP, Mfg date, Exp date",
     "Actual PM lot numbers, PM weights, Pack count per shift, Seal integrity check, Label verification (MRP and dates), Rejections count, Final packed qty, Operator name",
     "Formulation BMR (same batchNo), BOM Issuance (Packing), FG Inward",
     "YES — PDF"),
    ("Labeling / Finishing BMR",
     "Finishing / Labeling",
     "Product, Batch No, Label type, MRP, Mfg date, Exp date, Qty to label",
     "Label lot no., Rejected labels count, Verified by QA, Final labeled qty, Overprint verification",
     "Packing BMR, FG Inward",
     "YES"),
    ("SFG Storage Record",
     "All sections",
     "Product, Formulation Batch No, Qty stored (KG), Formulation date, Storage location",
     "Visual inspection result, Storage conditions (temp / humidity), Approval by QC",
     "Formulation BMR -> SFG table -> Future packing plan",
     "YES — referenced in future packing BMR"),
    ("Direct Packing from SFG",
     "Packing (using previously formulated SFG)",
     "SFG Batch No, Date of formulation, Qty available in SFG, Pack specs from original order — all pre-filled",
     "Qty taken from SFG for this batch, Balance SFG remaining, Packing actuals (same as Packing BMR)",
     "SFG stock record, Packing BMR, FG Inward",
     "YES — SFG batch details printed on BMR"),
]

r = 5
for row in BMR_TYPES:
    ws6.row_dimensions[r].height = 80
    bmr, plant, auto, op, linked, print_ = row
    cell(ws6, r, 1, bmr, bold=True, bg=AMBER, sz=10)
    cell(ws6, r, 2, plant, sz=9, italic=True, fg=MID_GREEN)
    cell(ws6, r, 3, auto, sz=9, wrap=True, bg=BLUE_LIGHT)
    cell(ws6, r, 4, op, sz=9, wrap=True)
    cell(ws6, r, 5, linked, sz=9, wrap=True, bg=LIGHT_GREEN, italic=True)
    cell(ws6, r, 6, print_, sz=9, h="center", bold=True)
    r += 1

ws6.row_dimensions[r].height = 10
r += 1
section_title(ws6, r, 1, "  THE SFG MEMORY LOOP — How the system remembers and reuses formulated material", span=6, bg=ACCENT_GREEN, fg=WHITE)
r += 1
ws6.row_dimensions[r].height = 110
c = ws6.cell(row=r, column=1,
             value=(
                 "STEP 1 — FORMULATE AND STORE:\n"
                 "Production formulates Batch B-001 of Aquasoft (50 KG). Instead of packing immediately, "
                 "the material is stored as SFG. System records: Product = Aquasoft (PROD-001), Batch = B-001, Qty = 50 KG, Formulation Date = 12-Jun-2025, Location = Rack-A3.\n\n"
                 "STEP 2 — NEW ORDER ARRIVES FOR SAME PRODUCT:\n"
                 "A new sales order comes for Aquasoft 20 KG. When the order is saved, system instantly checks the SFG table. "
                 "Finds 50 KG of Aquasoft (Batch B-001) in stock. "
                 "Planner receives alert on the plan card: 'SFG Available: 50 KG in Batch B-001 (formulated 12-Jun-2025). You can plan packing directly.'\n\n"
                 "STEP 3 — PLAN PACKING ONLY (SKIP FORMULATION):\n"
                 "Planner creates a PACKING-ONLY plan. Links to SFG Batch B-001. "
                 "System pre-fills packing BMR: SFG Batch B-001, Formulation date 12-Jun-2025, Qty 20 KG being used. "
                 "SFG balance auto-updated to 30 KG.\n\n"
                 "STEP 4 — DIGITAL BATCH SHEET WITH FULL TRACEABILITY:\n"
                 "Packing BMR prints with SFG batch details, formulation date, and all packing actuals. "
                 "Physical BMR binder has: SFG batch sheet (formulation) + Packing batch sheet. "
                 "Complete traceability from raw material -> formulation -> SFG -> packing -> dispatch."
             ))
c.fill = fill(AMBER)
c.font = font(size=9)
c.alignment = align(wrap=True, v="top")
c.border = border()
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — HOW BOM WORKS
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("7. How BOM Works")
set_col_widths(ws7, [24, 34, 32, 30])
ws7.freeze_panes = "A3"

banner(ws7, 1, "  SOM ERP — BOM (Bill of Materials) — 3 Types, How Each Works", 4)
ws7.row_dimensions[2].height = 6

BOM_BLOCKS = [
    ("WHAT IS A BOM?", MID_GREEN, [
        ("Definition",
         "A BOM lists every Raw Material needed to make 1 KG of a product. The system multiplies qty-per-unit by batch size to calculate total RM requirements. "
         "This drives both the production indent and the store issuance.",
         "Example — Aquasoft BOM (per 1 KG):",
         "Bacillus subtilis: 500 gm  |  Carrier (Talc): 450 gm  |  Wetting agent: 50 gm"),
        ("BOM Role Types",
         "Each BOM line has a roleType that tells the system which BOM type uses it:\n"
         "ACTIVE = the main microbial/botanical ingredient\n"
         "CARRIER = talc, kaolin, silica, etc.\n"
         "EXCIPIENT = wetting agents, stabilizers, binders\n"
         "PACKING = pouches, boxes, labels, tape, cartons",
         "Formulation BOM uses:", "ACTIVE + CARRIER + EXCIPIENT lines"),
        ("", "", "Packing BOM uses:", "PACKING lines only"),
    ]),
    ("TYPE 1 — FORMULATION BOM (With Recipe)", BLUE_DARK, [
        ("Trigger", "Planner clicks 'Send Formulation BOM' from the plan card in Planning module.",
         "System fetches:", "All ACTIVE + CARRIER + EXCIPIENT lines from RecipeDb where productCode matches"),
        ("RM Qty Calculation", "Each RM qty = qtyPerUnit (from BOM) x batchSize (KG) x noOfCycles.\n"
         "Example: Bacillus subtilis 0.5 KG/KG x 100 KG batch x 2 cycles = 100 KG required",
         "Indent created:", "Production Indent with each RM line and calculated qty"),
        ("Store Issues", "Store sees issuance task in BOM Issuance screen. Scans QR of each RM lot. Issues qty. Stock ledger debited.",
         "Traceability:", "Batch -> BOM -> RM Lot -> GRN -> Supplier (complete chain)"),
        ("BMR Link", "All issued lot numbers and actual quantities auto-populate the Formulation BMR for that batch.",
         "Digital record:", "BomSend -> BomIssuanceLine (rmCode, required qty, issued qty, lot no, date)"),
    ]),
    ("TYPE 2 — PACKING BOM (With Recipe)", PURPLE_DARK, [
        ("Trigger", "Planner clicks 'Send Packing BOM' from plan card — after formulation is done, or from a SFG-based plan.",
         "System fetches:", "All PACKING role lines from RecipeDb for the product"),
        ("PM Qty Calculation", "Each PM qty = qtyPerUnit x batchSize x cycles.\n"
         "Example: LD Pouch 1 KG: 1 unit/KG x 100 KG = 100 pouches  |  B-CBB box: 1 per 10 units = 10 boxes",
         "Indent:", "Packing material indent with all PM lines"),
        ("Store Issues", "Store issues packing materials against the BOM. QR scan per lot. Ledger debited. Packing BMR auto-fills.",
         "Record:", "BomSend (type=PACKING), BomIssuanceLine, StockLedger debit"),
    ]),
    ("TYPE 3 — DIRECT BOM ISSUANCE (No Recipe / RM Packing Direct)", ORANGE_DARK, [
        ("Use Case", "RMs or packing materials need to be issued without a formal product recipe.\n"
         "Examples: Repacking RM into smaller units, trial batches, common material top-ups, special order packaging.",
         "Access:", "Indent Management module -> New Indent -> Manual line entry (no BOM lookup)"),
        ("How It Works", "Planner opens Indent Management. Adds each RM/PM line manually: item code, description, qty, and purpose. "
         "No automatic BOM fetch. Quantities entered directly by planner based on their judgment.",
         "Store:", "Receives manual indent. Issues as normal. Stock ledger debited. Issuance record created."),
        ("Traceability", "Manual indent still creates full issuance records. Linked to a batch no. or marked as 'General Issue'.",
         "Audit trail:", "Raised by, items issued, batch linked, date, qty — all captured and reportable"),
    ]),
]

r = 3
for block_title, block_color, rows in BOM_BLOCKS:
    ws7.row_dimensions[r].height = 22
    section_title(ws7, r, 1, f"  {block_title}", span=4, bg=block_color, fg=WHITE, sz=11)
    r += 1
    for ci, h_text in enumerate(["Field", "Explanation", "Key Point", "Detail"], 1):
        hdr(ws7, r, ci, h_text, bg=block_color, fg=WHITE, sz=9)
    r += 1
    for row in rows:
        label, explanation, key, detail = row
        if label == "":
            ws7.row_dimensions[r].height = 22
        else:
            ws7.row_dimensions[r].height = 60
        cell(ws7, r, 1, label, bold=(label!=""), bg=HEADER_GRAY, sz=10)
        cell(ws7, r, 2, explanation, sz=9, wrap=True)
        cell(ws7, r, 3, key, sz=9, italic=True, fg=GRAY_TEXT, bg=AMBER)
        cell(ws7, r, 4, detail, sz=9, wrap=True, bg=LIGHT_GREEN)
        r += 1
    ws7.row_dimensions[r].height = 8
    r += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — STORE & QR SYSTEM
# ══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("8. Store & QR System")
set_col_widths(ws8, [24, 36, 28, 30])
ws8.freeze_panes = "A3"

banner(ws8, 1, "  SOM ERP — Store Operations, QR System & Multi-Item Location Plan", 4)
ws8.row_dimensions[2].height = 6

STORE_BLOCKS = [
    ("HOW THE QR STORE SYSTEM WORKS (CURRENT)", MID_GREEN, [
        ("QR Generation on GRN",
         "Every item received in GRN gets a unique QR code label printed. The QR encodes: item code, lot no., qty, supplier, GRN date.",
         "Scan action:", "Thermal printer at store prints QR label per lot on GRN approval"),
        ("Location Tagging",
         "Each physical storage location (rack, bin, shelf) has a permanent QR sticker. When RM is placed, "
         "store scans ITEM QR + LOCATION QR to link them in the system.",
         "Current rule:", "ONE item type per location (single-item constraint — changing to multi-item)"),
        ("Stock Issuance (QR Scan)",
         "Store scans the ITEM QR to issue. System checks lot balance. Deducts qty from ledger. Updates location.",
         "FEFO rule:", "First Expiry First Out — oldest lot issued first automatically"),
        ("Location Query",
         "Anyone can scan a location QR to see what is currently stored there: item, qty, expiry, last movement.",
         "Mobile / tablet:", "Scan -> instant location inventory view. No manual lookup needed."),
        ("Full Traceability",
         "Every QR scan creates a ledger entry with timestamp, user, and qty. Full chain: GRN -> Storage -> Issuance -> Production Batch.",
         "Audit:", "Who scanned, when, qty in/out — complete and immutable log"),
    ]),
    ("PLANNED CHANGE — MULTI-ITEM LOCATION STORAGE", ORANGE_DARK, [
        ("Current Limitation",
         "Currently each storage location holds ONE item type only. This is impractical when a bin physically stores multiple items together.",
         "Problem:", "Store creates workarounds or dummy locations when a bin holds 2+ items"),
        ("Planned Change",
         "Move to multi-item location model: each location can hold N different item lots. "
         "Instead of location -> item (1:1), it becomes location -> [item1, item2, item3] (1:N).",
         "New data model:", "LocationInventory table: locationId, itemCode, lotNo, qty, placedDate"),
        ("How Scanning Changes",
         "PLACING: Scan LOCATION QR -> Scan ITEM QR -> Enter qty. System adds to location list.\n"
         "PICKING: Scan LOCATION QR -> See all items stored -> Select item being issued.",
         "UI change:", "Location detail screen shows full inventory list (not just one item)"),
        ("Stock Query Improvement",
         "Scan any location QR -> See full list of all items, lots, qtys, and expiry dates stored there.",
         "Benefit:", "Store team instantly knows what is in Bin B3 without item-level search"),
        ("Database Change Required",
         "New table: location_inventory replacing the single item_code field on location records.",
         "Schema:", "location_inventory (id, location_id, item_code, lot_no, qty, placed_date, uom)"),
    ]),
    ("RM GRN INWARD FLOW", BLUE_DARK, [
        ("Receive at Gate",
         "RM arrives. Security logs vehicle. GRN raised in system: supplier, invoice no., PO reference, items, qty received.",
         "GRN record:", "Status = PENDING VERIFICATION until store confirms qty and quality"),
        ("QR Label Printing",
         "After GRN is verified and approved, QR labels auto-printed for each lot. Label shows: item name, code, lot, qty, expiry, GRN date.",
         "Thermal print:", "One QR label per lot. Stuck on the bag / carton / drum."),
        ("Storage and Tagging",
         "Store scans item QR + location QR. Item linked to bin/rack in system. Stock ledger credited.",
         "Ledger:", "StockLedger entry: item, qty, type=INWARD, lot, location, grnId, date"),
        ("Indent Satisfaction Check",
         "If an open production indent exists for this RM, system alerts planner and store that the indent can now be fulfilled.",
         "Notification:", "'Indent [X] can now be issued — [RM] received [Y KG].' Sent to Planner + Store."),
    ]),
    ("FG OUTWARD / DISPATCH FLOW", RED_DARK, [
        ("FG Pick for Dispatch",
         "Dispatch team selects sales order for dispatch. System shows FG lots available for the product and batch.",
         "FEFO:", "Oldest FG lot suggested first. Dispatch can select specific batch if needed."),
        ("QR Scan Outward",
         "Each FG carton / pack QR scanned during loading. System records serial, qty, vehicle, dispatch date.",
         "Ledger:", "StockLedger debit per pack scanned. Running FG balance updated live."),
        ("Gate Pass Generation",
         "Gate pass auto-generated with: vehicle no., driver, seal no., items, qty, customer, DI no.",
         "Print:", "Gate pass printed for physical record. Gate security retains copy."),
        ("SalesOrder Closure",
         "After QR scan outward complete, SalesOrder status auto-moves to DISPATCHED. Outward entry finalized.",
         "Chain:", "SalesOrder.status = DISPATCHED -> SalesOrderItem.status = DISPATCHED -> Outward record"),
    ]),
]

r = 3
for block_title, block_color, rows in STORE_BLOCKS:
    ws8.row_dimensions[r].height = 22
    section_title(ws8, r, 1, f"  {block_title}", span=4, bg=block_color, fg=WHITE, sz=11)
    r += 1
    for ci, h_text in enumerate(["Action / Step", "Description", "Key Point", "Detail / Rule"], 1):
        hdr(ws8, r, ci, h_text, bg=block_color, fg=WHITE, sz=9)
    r += 1
    for row in rows:
        ws8.row_dimensions[r].height = 65
        label, desc, key, detail = row
        cell(ws8, r, 1, label, bold=True, bg=HEADER_GRAY, sz=10)
        cell(ws8, r, 2, desc, sz=9, wrap=True)
        cell(ws8, r, 3, key, sz=9, italic=True, fg=GRAY_TEXT, bg=AMBER)
        cell(ws8, r, 4, detail, sz=9, wrap=True, bg=LIGHT_GREEN)
        r += 1
    ws8.row_dimensions[r].height = 8
    r += 1

# ── Tab Colors ────────────────────────────────────────────────────────────────
tab_colors = [DARK_GREEN, MID_GREEN, ACCENT_GREEN, AMBER_DARK,
              BLUE_DARK, ORANGE_DARK, PURPLE_DARK, RED_DARK]
for ws_, color in zip([ws, ws2, ws3, ws4, ws5, ws6, ws7, ws8], tab_colors):
    ws_.sheet_properties.tabColor = color

# ── Save ──────────────────────────────────────────────────────────────────────
out = "/sessions/peaceful-elegant-cerf/mnt/SOM-ERP-APP/SOM_ERP_MD_Overview.xlsx"
wb.save(out)
print(f"Saved: {out}")